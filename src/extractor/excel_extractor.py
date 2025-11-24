from openpyxl import load_workbook

from .extractor_base import ExtractorBase


class ExcelExtractor(ExtractorBase):
    def extract(self) -> str:
        wb = load_workbook(filename=self.path, read_only=True, data_only=True)
        try:
            lines = []
            for sn in wb.sheetnames:
                ws = wb[sn]
                lines.append(f"[{ws.title}]")
                for row in ws.iter_rows(values_only=True):
                    if all(not cell_value for cell_value in row):
                        continue
                    lines.append(
                        "\t".join([str(cell_value or "") for cell_value in row])
                    )
                lines.append("")
            return "\n".join(lines)
        finally:
            wb.close()
